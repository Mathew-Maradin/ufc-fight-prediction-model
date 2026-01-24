import re
from bs4 import BeautifulSoup
import time
import requests
import pandas as pd
import openpyxl
from halo import Halo
from spinners import Spinners
from openpyxl import load_workbook
from fractions import Fraction

# Create base excel files for data to populate
def create_excel_file() -> None:
    workbook_upcoming = openpyxl.Workbook()
    sheet = workbook_upcoming.active
    sheet['A1'] = 'R_fighter'
    sheet['B1'] = 'B_fighter'
    sheet['C1'] = 'Date'
    sheet['D1'] = 'Location'
    workbook_upcoming.save('../data/upcoming_fights.xlsx')

    workbook_completed = openpyxl.Workbook()
    sheet = workbook_completed.active
    sheet['A1'] = 'R_fighter'
    sheet['B1'] = 'B_fighter'
    sheet['C1'] = 'R_KD'
    sheet['D1'] = 'B_KD'
    sheet['E1'] = 'R_SIG_STR.'
    sheet['F1'] = 'B_SIG_STR.'
    sheet['G1'] = 'R_SIG_STR_pct'
    sheet['H1'] = 'B_SIG_STR_pct'
    sheet['I1'] = 'R_TOTAL_STR.'
    sheet['J1'] = 'B_TOTAL_STR.'
    sheet['K1'] = 'R_TD'
    sheet['L1'] = 'B_TD'
    sheet['M1'] = 'R_TD_pct'
    sheet['N1'] = 'B_TD_pct'
    sheet['O1'] = 'R_SUB_ATT'
    sheet['P1'] = 'B_SUB_ATT'
    sheet['Q1'] = 'R_REV'
    sheet['R1'] = 'B_REV'
    sheet['S1'] = 'R_CTRL'
    sheet['T1'] = 'B_CTRL'
    sheet['U1'] = 'R_HEAD'
    sheet['V1'] = 'B_HEAD'
    sheet['W1'] = 'R_BODY'
    sheet['X1'] = 'B_BODY'
    sheet['Y1'] = 'R_LEG'
    sheet['Z1'] = 'B_LEG'
    sheet['AA1'] = 'R_DISTANCE'
    sheet['AB1'] = 'B_DISTANCE'
    sheet['AC1'] = 'R_CLINCH'
    sheet['AD1'] = 'B_CLINCH'
    sheet['AE1'] = 'R_GROUND'
    sheet['AF1'] = 'B_GROUND'
    sheet['AG1'] = 'win_by'
    sheet['AH1'] = 'last_round'
    sheet['AI1'] = 'last_round_time'
    sheet['AJ1'] = 'Format'
    sheet['AK1'] = 'Referee'
    sheet['AL1'] = 'Details'
    sheet['AM1'] = 'date'
    sheet['AN1'] = 'location'
    sheet['AO1'] = 'Fight_type'
    sheet['AP1'] = 'Winner'
    workbook_completed.save('../data/completed_fights.xlsx')

# Extract links for past events and soonest upcoming event
def get_event_links() -> list[str,list[str]]:
    event_list = []
    url = "http://www.ufcstats.com/statistics/events/completed?page=all"

    res = requests.get(url)
    doc = BeautifulSoup(res.text, "html.parser")

    table = doc.find('table', { 'class': 'b-statistics__table-events' })
    events = table.find("tbody")
    cards = events.find_all("a")

    for card in cards:
        href = card.get('href')
        event_list.append(href)

    upcoming_event = event_list.pop(0)

    return [upcoming_event, event_list]

# Extract matchups for upcoming fight card
def get_upcoming_fights(link) -> list[list[str], list[str]]:
    fights = []

    response = requests.get(link)
    doc = BeautifulSoup(response.text, "html.parser")

    data = doc.find("tbody")
    fights = data.find_all("a", {"class": "b-link b-link_style_black"})

    current_fight = []
    for fight in fights:
        curr_fight = fight.text.strip()
        current_fight.append(curr_fight)

    paired_fights = []
    current_pair = []

    for name in current_fight:
        if name == "View Matchup":
            if current_pair:
                paired_fights.append(current_pair)
                current_pair = []
        else:
            current_pair.append(name)

  
    box = doc.find("ul")
    event_data = box.find_all('li', {"class":'b-list__box-list-item'})
    event = []

    for i in event_data:
        data = i.get_text(strip=True)
        event.append(data.split(":")[1])
    
    return [paired_fights, event]

def get_fight_links(event_url) -> list[list[str], list[str]]:
    fight_links = []

    response = requests.get(event_url)
    doc = BeautifulSoup(response.text, "html.parser")

    table = doc.find('table', { 'class': 'b-fight-details__table b-fight-details__table_style_margin-top b-fight-details__table_type_event-details js-fight-table' })
    fights = table.find("tbody")
    all_fights = fights.find_all("tr", { 'class': 'b-fight-details__table-row b-fight-details__table-row__hover js-fight-details-click' })
 
    for fight in all_fights:
        link = fight.get('data-link')
        fight_links.append(link)
    
    box = doc.find("ul")
    event_data = box.find_all('li', {"class":'b-list__box-list-item'})
    event = []

    for i in event_data:
        data = i.get_text(strip=True)
        event.append(data.split(":")[1])


    return [fight_links, event]

def get_fight_data(link, event_data):
    fight_data = []

    response = requests.get(link)
    doc = BeautifulSoup(response.text, "html.parser")

    tables = doc.find_all('table')

    try:
        total_table = tables[0]
    except:
        return []
    
    data = total_table.find_all("p", {"class": "b-fight-details__table-text"})
    for i in data:
        i = i.getText().strip()

        if "of" in i:
            frac = i.split(" of ")
            try: 
                fraction = float(frac[0])/float(frac[1])
            except:
                fraction = 0
            fight_data.append(fraction)
        elif "%" in i:
            fight_data.append(i.replace("%", ""))
        elif "-" in i:
            fight_data.append(0)
        elif ":" in i:
            
            time = i.split(":")
            digits = "".join(time)

            pos = 1
            seconds = 0

            for digit in reversed(digits):
                seconds += (pos * int(digit))

                if pos == 0:
                    pos = 10
                elif pos == 10:
                    pos = 60
                else: 
                    pos = 600
            
            fight_data.append(seconds)
        else:
            fight_data.append(i)

    specific_strikes_table = tables[2]
    strike_data = specific_strikes_table.find_all("p", {"class": "b-fight-details__table-text"})
    strike_data = strike_data[6:]
    for i in strike_data:
        i = i.getText().strip()
        if "of" in i:
            frac = i.split(" of ")
            try: 
                fraction = float(frac[0])/float(frac[1])
            except:
                fraction = 0
            fight_data.append(fraction)

        else:
            fight_data.append(i)

    time_pattern = re.compile(r'\b\d{1,2}:\d{2}\b') ###FIX DECISION DATE HERE
    bout_data = doc.find('div', {"class": "b-fight-details__content"})

    curr_data = []
    for string in bout_data.stripped_strings:
        if "of" in string:
            frac = i.split(" of ")
            fraction = float(frac[0])/float(frac[1])
            fight_data.append(fraction)

        elif ":" not in string or time_pattern.match(string):
            curr_data.append(string)
        
    tempdata = []
    if len(curr_data) > 6:
        tempdata = curr_data[:5]
        tempdata.append(tempdata[0])

    else:
        tempdata = curr_data

    fight_data = fight_data + tempdata

    fight_data.append(event_data[0])
    fight_data.append(event_data[1])

    fight_type = doc.find('i', {"class": "b-fight-details__fight-title"})
    for i in fight_type.stripped_strings:
        fight_data.append(i)
    
    try:
        fighter = doc.find('i', {"class": "b-fight-details__person-status b-fight-details__person-status_style_green"})
        fightname = fighter.find_next_sibling("div")
        fighter = fightname.find("a", {"class":"b-link b-fight-details__person-link"})
        fight_data.append(fighter.text)
    except:
        fight_data.append("")

    return fight_data

def add_data_to_excel(data, type) -> None:
    if type == "completed":
        wb = load_workbook("../data/completed_fights.xlsx")
        sheet = wb.active
        
        for row_data in data:
            sheet.append(row_data)

        wb.save("../data/completed_fights.xlsx")
    else:
        wb = load_workbook("../data/upcoming_fights.xlsx")
        sheet = wb.active
        
        for row_data in data:
            sheet.append(row_data)

        wb.save("../data/upcoming_fights.xlsx")

def fetch_fight_data(link_event_tuple):
    link, event_data = link_event_tuple
    print(event_data)
    try:
        data = get_fight_data(link, event_data)
        print(data)
        return data
    except Exception as e:
        print(f"Error fetching {link}: {e}")
        return []
    

if __name__ == "__main__":
    # ---------------CREATING EXCEL FILES------------------------ 
    start_time = time.time()
    spinner = Halo(text='Creating Excel Files', spinner='dots')
    spinner.start()
    create_excel_file() 
    spinner.stop() 

    # ---------------FETCHING EVENT LINKS------------------------ 
    spinner = Halo(text='Fetching links for each event!', spinner='dots')
    spinner.start()
    upcoming_event_link, past_event_links = get_event_links()
    spinner.stop()

    # ---------------UPCOMING FIGHT DATA------------------------ 
    spinner = Halo(text='Fetching upcoming fights', spinner='dots')
    spinner.start()
    upcoming_fights, event_data = get_upcoming_fights(upcoming_event_link)

    for i in range(len(upcoming_fights)):
        upcoming_fights[i].append(event_data[0])
        upcoming_fights[i].append(event_data[1])
        
    add_data_to_excel(upcoming_fights, "upcoming")
    spinner.stop()

    # ---------------EXISTING FIGHT DATA-------------------------- 
    spinner = Halo(text='Fetching past fight data', spinner='dots')
    spinner.start()

    bulk_fight_data = []
    for evnet_link in past_event_links:
        fight_links, event_data = get_fight_links(evnet_link)

        for link in fight_links:
            print(link)
            data = get_fight_data(link, event_data)
            bulk_fight_data.append(data)
            
            for info in data:
                print(info)
            break
        break

#     # add_data_to_excel(bulk_fight_data, "completed")

#     spinner.stop()
#     end_time = time.time()
#     elapsed_time = end_time - start_time
#     print(f"Elapsed time: {elapsed_time} seconds")
    
    print(bulk_fight_data)